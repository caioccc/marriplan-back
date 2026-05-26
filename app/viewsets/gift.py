import io
import secrets
import re
from decimal import Decimal
from urllib.parse import parse_qs, unquote, urlparse

import openpyxl
import pandas as pd
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.text import slugify
from django.db import transaction
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from rest_framework import (filters, permissions, status, viewsets)
from rest_framework.decorators import (action)
from rest_framework.generics import get_object_or_404
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import AllowAny
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.views import APIView

from app.models import (Gift, GiftListShareToken, ProductCatalog)
from app.serializers import (GiftListShareTokenSerializer, GiftSerializer)
from app.logging_utils import audit_log
from app.utils import (
    MAX_GIFTS_PER_WEDDING,
    is_limit_reached,
    is_valid_url,
    notify_gift_status_change,
    normalize_gift_value,
    to_sentence_case,
    to_upper_camel_words,
)


ESSENTIAL_CATALOG_TO_GIFT_CATEGORY = {
    'cozinha': 'kitchen',
    'eletrodomesticos': 'electronics',
    'moveis': 'furniture',
    'quarto': 'home',
    'banheiro': 'home',
    'decoracao': 'decor',
    'tecnologia': 'electronics',
    'lavanderia': 'home',
    'lua_de_mel': 'travel',
    'outros': 'other',
}


AMAZON_PRODUCT_PATTERN = re.compile(r'/(?:dp|gp/product|gp/aw/d|product)/([A-Z0-9]{10})(?:[/?]|$)', re.IGNORECASE)


def map_catalog_category_to_gift(category: str | None, title: str | None = None) -> str:
    raw_value = f'{category or ""} {title or ""}'.strip().lower()
    normalized = (category or '').strip().lower()

    if normalized in ESSENTIAL_CATALOG_TO_GIFT_CATEGORY:
        return ESSENTIAL_CATALOG_TO_GIFT_CATEGORY[normalized]

    heuristics = (
        ('kitchen', ('panela', 'cafeteira', 'liquidificador', 'frigideira', 'air fryer', 'cooktop', 'micro-ondas')),
        ('electronics', ('tv', 'notebook', 'tablet', 'smartphone', 'fone', 'caixa de som', 'aspirador', 'ventilador')),
        ('furniture', ('mesa', 'cadeira', 'sofa', 'sofá', 'guarda roupa', 'armario', 'armário', 'rack', 'cama')),
        ('decor', ('quadro', 'vaso', 'espelho', 'abajur', 'luminaria', 'luminária', 'vela', 'aromatizador')),
        ('travel', ('mala', 'mochila', 'necessaire', 'travesseiro de pescoco', 'power bank')),
    )

    for gift_category, keywords in heuristics:
        if any(keyword in raw_value for keyword in keywords):
            return gift_category

    return 'home'


def build_gift_product_code(*, name: str | None = None, link: str | None = None, category: str | None = None) -> str:
    source = str(link or name or category or '').strip().lower()
    if not source:
        return ''

    if link:
        for candidate in (source, unquote(source)):
            match = AMAZON_PRODUCT_PATTERN.search(candidate)
            if match:
                return f'amazon-asin-{match.group(1).lower()}'

        parsed = urlparse(source)
        query_params = parse_qs(parsed.query)
        for key in ('url', 'u', 'redirect', 'dest', 'destination'):
            target = query_params.get(key, [None])[0]
            if not target:
                continue
            nested_code = build_gift_product_code(link=unquote(target))
            if nested_code:
                return nested_code

        normalized = f'{parsed.netloc}{parsed.path}'.strip('/') or source
        normalized = slugify(normalized)
        return f'url-{normalized[:90]}' if normalized else ''

    normalized_name = slugify(source)
    return f'name-{normalized_name[:90]}' if normalized_name else ''


class GenerateBasicGiftListAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        wedding_profile = getattr(user, 'wedding_profile', None)
        if not wedding_profile:
            return Response({'detail': 'Usuário sem perfil de casamento.'}, status=status.HTTP_400_BAD_REQUEST)

        catalog_items = list(
            ProductCatalog.objects.filter(is_essential_template=True).order_by('store', 'category', 'title')
        )
        if not catalog_items:
            return Response(
                {'detail': 'Nenhum item essencial foi encontrado no catálogo.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        existing_codes = set(
            Gift.objects.filter(wedding_profile=wedding_profile)
            .exclude(product_code='')
            .values_list('product_code', flat=True)
        )
        existing_codes.update(
            filter(
                None,
                (
                    build_gift_product_code(
                        name=gift.name,
                        link=gift.link,
                        category=gift.category,
                    )
                    for gift in Gift.objects.filter(wedding_profile=wedding_profile)
                ),
            )
        )

        gifts_to_create = []
        now = timezone.now()
        for catalog_item in catalog_items:
            product_code = build_gift_product_code(
                name=catalog_item.title,
                link=catalog_item.product_url,
                category=catalog_item.category,
            )
            if not product_code:
                continue
            if product_code in existing_codes:
                continue

            gifts_to_create.append(
                Gift(
                    wedding_profile=wedding_profile,
                    name=catalog_item.title,
                    value=catalog_item.price or Decimal('0.00'),
                    link=catalog_item.product_url,
                    description=catalog_item.description,
                    category=map_catalog_category_to_gift(catalog_item.category, catalog_item.title),
                    image=catalog_item.image_url,
                    image_public_id='',
                    icon='',
                    status='available',
                    purchased_by='',
                    purchase_date=None,
                    reserved_by='',
                    reserved_message='',
                    reserved_at=None,
                    product_code=product_code,
                    created_at=now,
                    updated_at=now,
                )
            )
            existing_codes.add(product_code)

        if not gifts_to_create:
            return Response(
                {
                    'created': 0,
                    'skipped_existing': len(catalog_items),
                    'detail': 'A lista básica já foi gerada para este casamento.',
                },
                status=status.HTTP_200_OK,
            )

        current_count = Gift.objects.filter(wedding_profile=wedding_profile).count()
        if current_count + len(gifts_to_create) > MAX_GIFTS_PER_WEDDING:
            return Response(
                {
                    'detail': 'A lista básica excede o limite máximo de presentes permitido para este casamento.',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            Gift.objects.bulk_create(gifts_to_create, batch_size=100, ignore_conflicts=True)

        audit_log(
            'gift.generate_basic',
            user=request.user,
            obj=wedding_profile,
            message=f'Lista básica gerada com {len(gifts_to_create)} presente(s)',
        )
        return Response(
            {
                'created': len(gifts_to_create),
                'skipped_existing': len(catalog_items) - len(gifts_to_create),
                'detail': 'Lista básica de presentes gerada com sucesso.',
            },
            status=status.HTTP_201_CREATED,
        )


class GiftViewSet(viewsets.ModelViewSet):
    queryset = Gift.objects.all()
    serializer_class = GiftSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'description', 'category', 'product_code']
    ordering_fields = ['name', 'value', 'category', 'status', 'created_at']
    ordering = ['-created_at']

    def create(self, request, *args, **kwargs):
        user = request.user
        wedding_profile = getattr(user, 'wedding_profile', None)
        if not wedding_profile:
            return Response({'detail': 'Usuário sem perfil de casamento.'}, status=400)
        current_count = Gift.objects.filter(wedding_profile=wedding_profile).count()
        if is_limit_reached(current_count, MAX_GIFTS_PER_WEDDING):
            raise ValidationError({'detail': 'Limite de 300 presentes atingido.'})
        data = request.data.copy()
        data['wedding_profile'] = wedding_profile.id
        product_code = data.get('product_code') or build_gift_product_code(
            name=data.get('name'),
            link=data.get('link'),
            category=data.get('category'),
        )
        if product_code:
            existing_gift = Gift.objects.filter(
                wedding_profile=wedding_profile,
                product_code=product_code,
            ).first()
            if existing_gift:
                return Response(
                    GiftSerializer(existing_gift, context={'request': request}).data,
                    status=status.HTTP_200_OK,
                )

        data['product_code'] = product_code
        serializer = self.get_serializer(data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        # notify_gift_status_change(serializer.instance, 'created') # Notificação para created ainda não implementada
        audit_log('gift.create', user=request.user, obj=serializer.instance, message='Presente criado')
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        if request.data.get('status') == 'available':
            gift = self.get_object()
            if gift.status == 'available':
                gift.purchased_by = ''
                gift.purchase_date = None
                gift.reserved_by = ''
                gift.reserved_message = ''
                gift.reserved_at = None
                gift.save()
                response.data = GiftSerializer(gift, context={'request': request}).data
            audit_log('gift.update', user=request.user, obj=self.get_object(), message='Presente atualizado')
        return response

    def perform_destroy(self, instance):
        audit_log('gift.delete', user=self.request.user, obj=instance, message='Presente removido')
        instance.delete()

    def get_queryset(self):
        user = self.request.user
        wedding_profile = getattr(user, 'wedding_profile', None)
        qs = Gift.objects.all()
        if wedding_profile:
            qs = qs.filter(wedding_profile=wedding_profile)
        status_param = self.request.query_params.get('status')
        if status_param:
            qs = qs.filter(status=status_param)
        category_param = self.request.query_params.get('category')
        if category_param:
            qs = qs.filter(category=category_param)
        return qs

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def mark_as_purchased(self, request, pk=None):
        gift = self.get_object()
        if gift.status == 'purchased':
            return Response({'detail': 'Gift already marked as purchased.'}, status=400)
        gift.status = 'purchased'
        gift.purchased_by = request.data.get('purchased_by',
                                             '') or request.user.get_full_name() or request.user.username
        gift.purchase_date = timezone.now()
        gift.reserved_by = ''
        gift.reserved_message = ''
        gift.reserved_at = None
        gift.save()
        message = request.data.get('message', '')
        notify_gift_status_change(gift, 'purchased', message=message)
        return Response(GiftSerializer(gift, context={'request': request}).data)

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def unmark_as_purchased(self, request, pk=None):
        gift = self.get_object()
        if gift.status != 'purchased':
            return Response({'detail': 'Gift is not marked as purchased.'}, status=400)
        gift.status = 'available'  # ou o valor default do status
        gift.purchased_by = ''
        gift.purchase_date = None
        gift.reserved_by = ''
        gift.reserved_message = ''
        gift.reserved_at = None
        gift.save()
        notify_gift_status_change(gift, 'unmarked')
        return Response(GiftSerializer(gift, context={'request': request}).data)

    @action(detail=False, methods=['get'], url_path='template')
    def download_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Cabeçalhos em português (sem Ícone)
        ws.append([
            'Nome do Presente',
            'Valor',
            'Link',
            'Descrição',
            'Categoria',
            'Imagem',
            'Status',
            'Código do Produto'
        ])
        # 3 exemplos realistas
        ws.append([
            'Panela Elétrica de Arroz',
            199.90,
            'https://www.amazon.com.br/panela-eletrica-de-arroz',
            'Panela elétrica para preparar arroz de forma prática e rápida.',
            'Casa',
            'https://images.unsplash.com/photo-1519864600265-abb23847ef2c',
            'Disponível',
            'PA-123'
        ])
        ws.append([
            'Jogo de Toalhas 5 peças',
            149.99,
            '',
            'Jogo de toalhas 100% algodão, super macias.',
            'Casa',
            'https://images.unsplash.com/photo-1506744038136-46273834b3fb',
            'Disponível',
            'JT-456'
        ])
        ws.append([
            'Liquidificador Turbo',
            249.00,
            'https://www.amazon.com.br/liquidificador-turbo',
            'Liquidificador potente com 5 velocidades.',
            'Eletrodomésticos',
            'https://images.unsplash.com/photo-1519125323398-675f0ddb6308',
            'Disponível',
            'LQ-789'
        ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=modelo_lista_presentes.xlsx'
        return response

    @action(detail=False, methods=['post'], url_path='import', parser_classes=[MultiPartParser])
    def import_gifts(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Arquivo não enviado.'}, status=400)
        # Mapeamento de cabeçalhos PT-BR para campos do model (sem Ícone)
        header_map = {
            'Nome do Presente': 'name',
            'Valor': 'value',
            'Link': 'link',
            'Descrição': 'description',
            'Categoria': 'category',
            'Imagem': 'image',
            'Imagem Public ID': 'image_public_id',
            'Status': 'status',
            'Código do Produto': 'product_code',
        }
        fields = list(header_map.values())
        # Funções de normalização movidas para cá
        def normalize_status(val):
            if not val:
                return 'available'
            for code, label in Gift.STATUS_CHOICES:
                if val == code or val.lower() == label.lower() or val.lower() == label.lower().capitalize():
                    return code
            return 'available'
        def normalize_category(val):
            if not val:
                return 'other'
            for code, label in Gift.CATEGORY_CHOICES:
                if val == code or val.lower() == label.lower() or val.lower() == label.lower().capitalize():
                    return code
            return 'other'
        def validate_row(data, row_idx):
            errors = []
            if not data.get('name') or str(data.get('name')).strip() == '':
                errors.append(f"Linha {row_idx}: Nome do presente é obrigatório.")
            if not data.get('value') or str(data.get('value')).strip() == '':
                errors.append(f"Linha {row_idx}: Valor é obrigatório.")
            if data.get('image') and not is_valid_url(data.get('image')):
                errors.append(f"Linha {row_idx}: Imagem deve ser um link válido (URL).")
            # Choices válidos do model
            VALID_STATUS = [c[0] for c in Gift.STATUS_CHOICES] + [c[1] for c in Gift.STATUS_CHOICES]
            VALID_CATEGORY = [c[0] for c in Gift.CATEGORY_CHOICES] + [c[1] for c in Gift.CATEGORY_CHOICES]
            status_value = data.get('status')
            if status_value and status_value not in VALID_STATUS:
                errors.append(f"Linha {row_idx}: Status inválido.")
            category_value = data.get('category')
            if category_value and category_value not in VALID_CATEGORY:
                errors.append(f"Linha {row_idx}: Categoria inválida.")
            return errors

        def normalize_row(data):
            data['name'] = to_upper_camel_words(data.get('name'))
            data['description'] = to_sentence_case(data.get('description'))
            data['value'] = normalize_gift_value(data.get('value'))

            if data.get('link') and not is_valid_url(data.get('link')):
                data['link'] = ''
            elif data.get('link'):
                data['link'] = str(data.get('link')).strip()

            if data.get('image') and not is_valid_url(data.get('image')):
                data['image'] = ''
            elif data.get('image'):
                data['image'] = str(data.get('image')).strip()

            if not data.get('product_code'):
                data['product_code'] = build_gift_product_code(
                    name=data.get('name'),
                    link=data.get('link'),
                    category=data.get('category'),
                )

            return data

        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
            if any(col in header_map for col in df.columns):
                df = df.rename(columns=header_map)
            created, errors = 0, []
            current_count = Gift.objects.filter(wedding_profile=request.user.wedding_profile).count()
            for i, row in df.iterrows():
                if is_limit_reached(current_count, MAX_GIFTS_PER_WEDDING):
                    errors.append(f'Linha {i + 2}: Limite de 300 presentes atingido. Os registros restantes foram ignorados.')
                    break
                try:
                    data = dict(zip(fields, [row.get(f, '') for f in fields]))
                    data['wedding_profile'] = request.user.wedding_profile.id
                    # Normalização e validação
                    data = normalize_row(data)
                    data['status'] = normalize_status(data.get('status'))
                    data['category'] = normalize_category(data.get('category'))
                    # Validações
                    row_errors = validate_row(data, i + 2)
                    if row_errors:
                        errors.extend(row_errors)
                        continue
                    for field in fields:
                        if field in data and (data[field] is None or str(data[field]).strip() == '' or pd.isna(data[field])):
                            data[field] = ''
                    serializer = self.get_serializer(data=data)
                    serializer.is_valid(raise_exception=True)
                    serializer.save()
                    created += 1
                    current_count += 1
                except Exception as e:
                    errors.append(f"Linha {i + 2}: {str(e)}")
        elif file.name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            mapped_headers = [header_map.get(h, h) for h in header_row]
            created, errors = 0, []
            current_count = Gift.objects.filter(wedding_profile=request.user.wedding_profile).count()
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if is_limit_reached(current_count, MAX_GIFTS_PER_WEDDING):
                    errors.append(f'Linha {i}: Limite de 300 presentes atingido. Os registros restantes foram ignorados.')
                    break
                try:
                    data = dict(zip(mapped_headers, row))
                    data = {f: data.get(f, '') for f in fields}
                    data['wedding_profile'] = request.user.wedding_profile.id
                    # Normalização e validação
                    data = normalize_row(data)
                    data['status'] = normalize_status(data.get('status'))
                    data['category'] = normalize_category(data.get('category'))
                    # Validações
                    row_errors = validate_row(data, i)
                    if row_errors:
                        errors.extend(row_errors)
                        continue
                    for field in fields:
                        if field in data and (data[field] is None or str(data[field]).strip() == '' or pd.isna(data[field])):
                            data[field] = ''
                    serializer = self.get_serializer(data=data)
                    serializer.is_valid(raise_exception=True)
                    serializer.save()
                    created += 1
                    current_count += 1
                except Exception as e:
                    errors.append(f"Linha {i}: {str(e)}")
        else:
            return Response({'detail': 'Formato não suportado. Envie CSV ou Excel.'}, status=400)

        return Response({'created': created, 'errors': errors})

    @action(detail=False, methods=['get'], url_path='export/pdf')
    def export_pdf(self, request):
        user = request.user
        wedding_profile = getattr(user, 'wedding_profile', None)
        if not wedding_profile:
            return Response({'detail': 'Usuário sem perfil de casamento.'}, status=400)
        queryset = Gift.objects.filter(wedding_profile=wedding_profile)
        data = GiftSerializer(queryset, many=True).data
        df = pd.DataFrame(data)
        try:
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="presentes.pdf"'
            buffer = io.BytesIO()
            p = canvas.Canvas(buffer, pagesize=letter)
            width, height = letter
            y = height - 40
            p.setFont('Helvetica-Bold', 14)
            p.drawString(40, y, 'Lista de Presentes')
            y -= 30
            p.setFont('Helvetica', 10)
            columns = ['Nome', 'Valor', 'Categoria', 'Status', 'Descrição']
            col_widths = [140, 80, 80, 60, 140]
            x = 40
            for i, col in enumerate(columns):
                p.drawString(x, y, col)
                x += col_widths[i]
            y -= 18
            for _, row in df.iterrows():
                x = 40
                values = [
                    str(row.get('name', '')),
                    str(row.get('value', '')),
                    str(row.get('category', '')),
                    str(row.get('status', '')),
                    str(row.get('description', ''))
                ]
                for i, value in enumerate(values):
                    p.drawString(x, y,
                                 value[:40] if i == 0 else value[:25])  # Permite nomes maiores na primeira coluna
                    x += col_widths[i]
                y -= 15
                if y < 60:
                    p.showPage()
                    y = height - 40
                    p.setFont('Helvetica', 10)
            p.save()
            pdf = buffer.getvalue()
            buffer.close()
            response.write(pdf)
            return response
        except Exception as e:
            return Response({'detail': f'Erro ao exportar PDF: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='all', permission_classes=[permissions.IsAuthenticated])
    def all(self, request):
        """Retorna todos os presentes do perfil do casal sem paginação (uso para dashboards/exports)."""
        user = request.user
        wedding_profile = getattr(user, 'wedding_profile', None)
        if not wedding_profile:
            return Response({'detail': 'Usuário sem perfil de casamento.'}, status=400)
        queryset = Gift.objects.filter(wedding_profile=wedding_profile)
        # Permitir filtros básicos por search/ordering/status/category se desejado
        search = request.GET.get('search')
        if search:
            queryset = queryset.filter(name__icontains=search) | queryset.filter(category__icontains=search)
        status_param = request.GET.get('status')
        if status_param:
            status_list = [s.strip() for s in status_param.split(',') if s.strip()]
            queryset = queryset.filter(status__in=status_list)
        category_param = request.GET.get('category')
        if category_param:
            queryset = queryset.filter(category=category_param)

        serializer = GiftSerializer(queryset, many=True, context={'request': request})
        return Response(serializer.data)

class GiftListShareTokenView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        wedding_profile = getattr(user, 'wedding_profile', None)
        if not wedding_profile:
            return Response({'detail': 'Usuário sem perfil de casamento.'}, status=400)
        token_obj, created = GiftListShareToken.objects.get_or_create(wedding_profile=wedding_profile)
        if created or not token_obj.token:
            token_obj.token = secrets.token_urlsafe(16)
            token_obj.save()

        return Response(GiftListShareTokenSerializer(token_obj).data)


class PublicGiftListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, token):
        token_obj = get_object_or_404(GiftListShareToken, token=token)
        wedding_profile = token_obj.wedding_profile
        gifts = Gift.objects.filter(wedding_profile=token_obj.wedding_profile)

        # Filtros múltiplos por categoria
        categories = request.GET.get('categories')
        if categories:
            cat_list = [c.strip() for c in categories.split(',') if c.strip()]
            gifts = gifts.filter(category__in=cat_list)

        # Filtro por faixa de preço
        min_price = request.GET.get('min_price')
        max_price = request.GET.get('max_price')
        if min_price:
            gifts = gifts.filter(value__gte=float(min_price))
        if max_price:
            gifts = gifts.filter(value__lte=float(max_price))

        # Filtro por status múltiplo
        status_param = request.GET.get('status')
        if status_param:
            status_list = [s.strip() for s in status_param.split(',') if s.strip()]
            gifts = gifts.filter(status__in=status_list)

        # Filtro por link
        has_link = request.GET.get('has_link')
        if has_link == 'true':
            gifts = gifts.exclude(link__isnull=True).exclude(link='')
        elif has_link == 'false':
            gifts = gifts.filter(link__isnull=True) | gifts.filter(link='')

        # Busca por nome/categoria
        search = request.GET.get('search')
        if search:
            gifts = gifts.filter(
                name__icontains=search
            ) | gifts.filter(
                category__icontains=search
            )

        # Ordenação
        ordering = request.GET.get('ordering', 'recent')
        if ordering == 'recent':
            gifts = gifts.order_by('-created_at')
        elif ordering == 'oldest':
            gifts = gifts.order_by('created_at')
        elif ordering == 'price_asc':
            gifts = gifts.order_by('value')
        elif ordering == 'price_desc':
            gifts = gifts.order_by('-value')

        # Paginação
        try:
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 12))
        except Exception:
            page, page_size = 1, 12
        total = gifts.count()
        start = (page - 1) * page_size
        end = start + page_size
        gifts_page = gifts[start:end]

        serializer = GiftSerializer(gifts_page, many=True)

        # Opções para filtros (categorias e status únicos)
        all_gifts = Gift.objects.filter(wedding_profile=token_obj.wedding_profile)
        categories_options = list(all_gifts.values_list('category', flat=True).distinct())
        status_options = list(all_gifts.values_list('status', flat=True).distinct())

        return Response({
            'results': serializer.data,
            'total': total,
            'page': page,
            'page_size': page_size,
            'categories': categories_options,
            'status': status_options,
            'wedding_profile': {
                'id': wedding_profile.id,
                'nome_noivo': wedding_profile.nome_noivo,
                'nome_noiva': wedding_profile.nome_noiva,
            },
        })

    def post(self, request, token):
        token_obj = get_object_or_404(GiftListShareToken, token=token)
        gift_id = request.data.get('gift_id')
        if not gift_id:
            return Response({'detail': 'Presente não informado.'}, status=status.HTTP_400_BAD_REQUEST)

        gift = get_object_or_404(Gift, pk=gift_id, wedding_profile=token_obj.wedding_profile)
        if gift.status != 'available':
            return Response({'detail': 'Este presente não está disponível para reserva.'}, status=status.HTTP_400_BAD_REQUEST)

        reserver_name = (request.data.get('reserver_name') or request.data.get('name') or '').strip()
        reservation_message = (request.data.get('message') or '').strip()

        with transaction.atomic():
            gift.status = 'reserved'
            gift.reserved_by = reserver_name
            gift.reserved_message = reservation_message
            gift.reserved_at = timezone.now()
            gift.purchased_by = ''
            gift.purchase_date = None
            gift.save()

        notification_result = notify_gift_status_change(
            gift,
            'reserved',
            message=reservation_message,
            reserved_by=reserver_name,
        )

        return Response({
            'gift': GiftSerializer(gift, context={'request': request}).data,
            'whatsapp_links': notification_result.get('whatsapp_links', []),
        }, status=status.HTTP_200_OK)
